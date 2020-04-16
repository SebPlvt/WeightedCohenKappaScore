# -*- coding: utf-8 -*-
"""
Created on Mon Sep  2 11:35:06 2019

@author: Sébastien Polvent
@version: 1.0

"""

import os
import openpyxl as pxl
import pandas as pd
import krippendorff


#TODO convert into python 3.7 data class
class ExcelRangeType():
    """ExcelRangeType"""
    def __init__(self, start_line, end_line, column):
        """ init """
        self.start_line = start_line
        self.column = column
        self.end_line = end_line



class WeightedCohenKappaScore():
    """ WeightedCohenKappaScore object"""


    def __init__(self):
        """ init """
        os.chdir(os.getcwd())
        self.scorer1_scoring = pd.DataFrame()
        self.scorer2_scoring = pd.DataFrame()
        self.df_results = pd.DataFrame()

        self.df_global = pd.DataFrame()

        self.scorer1_name = 'MH'
        self.scorer2_name = 'MB'

        self.xl_range = ExcelRangeType(start_line=6, end_line=277, column=3)

        self.scorer2_file = 'OSS_MB_relecture.xlsx'

        self.scoring_items = ['W', 'N1', 'N2', 'N3', 'R'] #['0', '1', '2', '3', '4']
        #liste_test = ['item' + str(x) for x in range(len(self.scoring_items)+1)]
        #matrix_test, inverse_matrix_test = create_coeff_matrixes(liste=liste_test)

        #linear coefficients matrix
        self.matrix_df_linear, self.inverse_matrix_df_linear = \
            self.create_coeff_matrices(items=self.scoring_items, \
                                       weights='linear')
        #quadratic coefficents matrix.
        self.matrix_df_square, self.inverse_matrix_df_square = \
            self.create_coeff_matrices(items=self.scoring_items, \
                                       weights='quadratic')

        self.result_filename = 'OSS_Agreement.xlsx'




    def import_scoring_from_1_file(self):
        """import_scoring_from_1_file_no_excel"""
        if self.xl_range and self.scorer2_file:
            #retrieve xl sheets names
            scoring_file = pxl.load_workbook(self.scorer2_file)
            sheet_names = []
            columns_names = []
            for sht in scoring_file:
                sheet_names.append(sht.title)
                columns_names.append(str(sht.title).replace(' ', '_'))
            scoring_file.close()
            #read data from sheets with pandas
            for sht, col_name in zip(sheet_names, columns_names):
                if not str(sht).count('Feuille'):
                    sht_df = pd.read_excel(io=self.scorer2_file, \
                                           sheet_name=sht, skiprows=3, header=1, \
                                           usecols=[1, 2])
                    if not pd.isna(sht_df.iloc[0, 0]):   #(6, 2)
                        self.scorer2_scoring[col_name] = \
                            sht_df[sht_df.columns[1]].dropna()
            #sort dataframe columns by name
            self.scorer2_scoring = \
                self._sort_dataframe(dataframe=self.scorer2_scoring)



    def import_scoring_from_many_files(self):
        """import_scoring_from_many_files_no_excel"""
        if self.xl_range:
            all_files_list = os.listdir(os.getcwd())
            xlsx_files_list = []
            #keep only xlsx files
            for file in all_files_list:
                if file.count('.xlsx') \
                and not file.count(self.scorer2_file[:len(self.scorer2_file)-5]) \
                and not file.count('OSS_Agreement'):
                    xlsx_files_list.append(file)
            for xl_file in xlsx_files_list:
                sht_df = pd.read_excel(io=xl_file, sheet_name=0, skiprows=3, \
                                       header=1, usecols=[1, 2])
                if not pd.isna(sht_df.iloc[0, 0]):   #(6, 2)
                    self.scorer1_scoring[xl_file[4:len(xl_file)-5]] = \
                        sht_df[sht_df.columns[1]].dropna()
            #sort dataframe columns by name
            self.scorer1_scoring = \
                self._sort_dataframe(dataframe=self.scorer1_scoring)



    def _sort_dataframe(self, dataframe):
        """sort dataframe columns by name"""
        columns = list(dataframe.columns)
        columns.sort()
        dataframe = dataframe[columns]
        return dataframe



    def create_coeff_matrices(self, items, weights='linear'):
        """ x*x matrice coeffs and inverse
        item : scoring items list
        weights : str, optional,
        Weighting type to calculate the score.
        “linear” means linear weighted; “quadratic” means quadratic weighted.
        default : "linear"
        Returns 2 matrices :
            mtx : advantage to agreement
            inv_mtx : disadvantage biggest errors
        """
        mtx = pd.DataFrame(index=items, columns=items)
        for col in range(len(mtx.columns)):
            for idx in range(len(mtx.index)):
                if col == idx:
                    mtx.iloc[idx, col] = 1
                else:
                    temp_value = abs(idx - col) / (len(items) - 1)
                    if weights == 'linear':
                        mtx.iloc[idx, col] = 1 - temp_value
                    elif weights == 'quadratic':
                        mtx.iloc[idx, col] = 1 - temp_value ** 2
        inv_mtx = abs(1 - mtx)
        return mtx, inv_mtx



    def compute_kappa_score(self, scorer1, scorer2, items, matrix_weights):
        """computes and returns weighted kappa score from 2 lists or pandas series
        of scorings.
        Compare results wiht scikit-learn : sklearn.metrics.cohen_kappa_score

        #Idem p_a
        p_a = 0
        for lig in df_n_items.index:
            for col in df_n_items.columns:
                p_a = p_a + (df_n_items.loc[lig, col] / len(scorer1)) \
                      * matrix_weights.loc[lig, col]"""
        #matrice nb items
        df_n_items = pd.DataFrame(data=0, index=items, columns=items)
        for score1, score2 in zip(scorer1, scorer2):
            df_n_items[score1][score2] = df_n_items[score1][score2] + 1
        #nb items * total number of scoring items
        df_proba = df_n_items / len(scorer1)
        #df_proba with weights
        df_proba_w = df_proba * matrix_weights
        #relative observed agreement among raters
        p_a = df_proba_w.sum().sum()
        #hypothetical probability of chance agreement
        p_e = 0
        for lig in df_proba.index:
            for col in df_proba.columns:
                p_e = p_e + df_proba.loc[lig, :].sum() * \
                    df_proba.loc[:, col].sum() * matrix_weights.loc[lig, col]
        #kappa score
        kappa = 1 - (1 - p_a) / (1 - p_e)
        return kappa



    def compute_kappa(self):
        """Prepares data for Cohen's kappa score computation,
            score results and corresponding indexes  """
        kappa_res_linear = []
        kappa_res_inv_linear = []
        kappa_res_quadratic = []
        kappa_res_inv_quadratic = []
        kappa_idx = []
        for score1 in  self.scorer1_scoring.columns:
            for score2 in self.scorer2_scoring.columns:
                if score1 == score2:
                    kappa_res_linear.append(
                        self.compute_kappa_score(scorer1=self.scorer1_scoring[score1], \
                                                 scorer2=self.scorer2_scoring[score2], \
                                                 items=self.scoring_items, \
                                                 matrix_weights=self.matrix_df_linear))
                    kappa_res_inv_linear.append(
                        self.compute_kappa_score(scorer1=self.scorer1_scoring[score1], \
                                                 scorer2=self.scorer2_scoring[score2], \
                                                 items=self.scoring_items, \
                                                 matrix_weights=self.inverse_matrix_df_linear))
                    kappa_res_quadratic.append(
                        self.compute_kappa_score(scorer1=self.scorer1_scoring[score1], \
                                                 scorer2=self.scorer2_scoring[score2], \
                                                 items=self.scoring_items, \
                                                 matrix_weights=self.matrix_df_square))
                    kappa_res_inv_quadratic.append(
                        self.compute_kappa_score(scorer1=self.scorer1_scoring[score1], \
                                                 scorer2=self.scorer2_scoring[score2], \
                                                 items=self.scoring_items, \
                                                 matrix_weights=self.inverse_matrix_df_square))
                    if self.df_results.index.empty:
                        kappa_idx.append(score1)
                    break
        if self.df_results.empty:
            self.df_results = pd.DataFrame(index=kappa_idx)

        self.df_results["Cohen's kappa score (linear coefficients)"] = \
            kappa_res_linear
        self.df_results["Cohen's kappa score (linear inverse coefficients)"] = \
            kappa_res_inv_linear
        self.df_results["Cohen's kappa score (quadratic coefficients)"] = \
            kappa_res_quadratic
        self.df_results["Cohen's kappa score (quadratic inverse coefficients)"] = \
            kappa_res_inv_quadratic

        self.df_results['Kappa Agreement linear'] = \
            self.kappa_interpretation(
                self.df_results["Cohen's kappa score (linear coefficients)"])

        self.df_results['Kappa Agreement linear inverse'] = \
            self.kappa_interpretation(
                self.df_results["Cohen's kappa score (linear inverse coefficients)"])

        self.df_results['Kappa Agreement quadratic'] = \
            self.kappa_interpretation(
                self.df_results["Cohen's kappa score (quadratic coefficients)"])

        self.df_results['Kappa Agreement quadratic inverse'] = \
            self.kappa_interpretation(
                self.df_results["Cohen's kappa score (quadratic inverse coefficients)"])



    def kappa_interpretation(self, serie):
        """kappa score interpretation"""
        kappa_interpretation = []
        for kappa in serie:
            if kappa < 0:
                kappa_interpretation.append('Less than chance agreeement')
            elif 0 <= kappa < 0.21:
                kappa_interpretation.append('Slight agreement')
            elif 0.20 < kappa < 0.41:
                kappa_interpretation.append('Fair agreement')
            elif 0.40 < kappa < 0.61:
                kappa_interpretation.append('Moderate agreement')
            elif 0.60 < kappa < 0.81:
                kappa_interpretation.append('Substantial agreement')
            elif 0.80 < kappa <= 1:
                kappa_interpretation.append('Almost perfect agreement')
            elif kappa == 1:
                kappa_interpretation.append('Perfect agreement')
        return kappa_interpretation



    def compute_global_kappa(self):
        """compute_global_kappa"""
        drop_ok = self._drop_single_columns()
        if drop_ok:
            #global kappa
            global_kappa_score = pd.Series()
            global1 = []
            for col in self.scorer1_scoring.columns:
                global1.extend(self.scorer1_scoring[col].tolist())
            global2 = []
            for col in self.scorer2_scoring.columns:
                global2.extend(self.scorer2_scoring[col].tolist())
            global_kappa_score['Global kappa score (linear coefficients)'] = \
                self.compute_kappa_score(scorer1=global1, scorer2=global2, \
                                         items=self.scoring_items, \
                                         matrix_weights=self.matrix_df_linear)
            global_kappa_score['Global kappa score (quadratic coefficients)'] = \
                self.compute_kappa_score(scorer1=global1, scorer2=global2, \
                                         items=self.scoring_items, \
                                         matrix_weights=self.matrix_df_square)

            global_kappa_score['Global kappa score (linear inverse coefficients)'] = \
                self.compute_kappa_score(scorer1=global1, scorer2=global2, \
                                         items=self.scoring_items, \
                                         matrix_weights=self.inverse_matrix_df_linear)
            global_kappa_score['Global kappa score (quadratic inverse coefficients)'] = \
                self.compute_kappa_score(scorer1=global1, scorer2=global2, \
                                         items=self.scoring_items, \
                                         matrix_weights=self.inverse_matrix_df_square)

            self.df_global['Global Kappa Scores'] = global_kappa_score
            self.df_global['Kappa Interpretation'] = \
                self.kappa_interpretation(global_kappa_score)
        else:
            print('Error while dropping single columns !\n' + \
                  'Global Kappa computing aborted.')



    def compute_krippendorff_alpha(self, level_of_measurement='ordinal'):
        """compute_krippendorff_alpha"""
        kri_res = []
        kri_idx = []
        for scorer1 in  self.scorer1_scoring.columns:
            for scorer2 in self.scorer2_scoring.columns:
                if scorer1 == scorer2:
                    comp_df = pd.DataFrame(columns=[self.scorer1_name, \
                                                    self.scorer2_name])
                    comp_df[self.scorer1_name] = self.scorer1_scoring[scorer1]
                    comp_df[self.scorer2_name] = self.scorer2_scoring[scorer2]
                    comp_df = comp_df.T
                    kri_res.append(krippendorff.alpha(reliability_data=comp_df, \
                                                      value_domain=self.scoring_items, \
                                                      level_of_measurement=level_of_measurement))
                    if self.df_results.index.empty:
                        kri_idx.append(scorer1)
                    break
        self.df_results["Krippendorff alpha score"] = kri_res
        self.df_results["Alpha score Agreement"] = \
            self.alpha_interpretation(self.df_results["Krippendorff alpha score"])



    def alpha_interpretation(self, serie):
        """kripendorff alpha interpretation
           NOT SURE OF THE SCALE ! """
        alpha_interpretation = []
        for alpha in serie:
            if alpha <= 0.667:
                alpha_interpretation.append('Unreliable agreement')
            elif 0.667 < alpha < 0.81:
                alpha_interpretation.append('Acceptable agreement')
            elif 0.80 < alpha <= 1:
                alpha_interpretation.append('Substantial agreement')
            elif alpha == 1:
                alpha_interpretation.append('Perfect agreement')
        return alpha_interpretation



    def save_results_to_excel(self):
        """save_results_to_excel"""
        #test if a previous result file already exists
        if os.path.exists('./' + self.result_filename):
            #test if old result file already exists, then delete it
            if os.path.exists('./' + self.result_filename[:-5] + '_old.xlsx'):
                os.remove('./' + self.result_filename[:-5] + '_old.xlsx')
            #rename previous result file
            os.rename('./' + self.result_filename, \
                      './' + self.result_filename[:-5] + '_old.xlsx')
        #copy data to xl
        with pd.ExcelWriter(self.result_filename) as writer:
            self.df_results.to_excel(excel_writer=writer, \
                                     sheet_name='Agreement')
            self.df_global.to_excel(excel_writer=writer, \
                                    sheet_name='Global Agreement')



    def _drop_single_columns(self):
        """drop_single_columns"""
        #drop single scored columns
        scorer1_len = len(self.scorer1_scoring.columns)
        scorer2_len = len(self.scorer2_scoring.columns)
        if scorer1_len > scorer2_len:
            for item in self.scorer1_scoring:
                if item not in self.scorer2_scoring.columns:
                    del self.scorer1_scoring[item]
        scorer1_len = len(self.scorer1_scoring.columns)
        if scorer2_len > scorer1_len:
            for item in self.scorer2_scoring:
                if item not in self.scorer1_scoring.columns:
                    del self.scorer2_scoring[item]
        scorer2_len = len(self.scorer2_scoring.columns)
        return scorer1_len == scorer2_len



    def highlight_differences_scorer2(self):
        """highlight_differences_scorer2_no_excel"""
        drop_ok = self._drop_single_columns()
        if drop_ok:
            #compare each scoring item
            scoring_difference = self.scorer1_scoring != self.scorer2_scoring
            #highlight differences in scorer2's scoring file and add scorer1's scoring in col E
            scorer2_xl = pxl.load_workbook(self.scorer2_file)
            # bg color RGB(255, 0, 0) hex '#ff0000'
            bg_style = pxl.styles.PatternFill(start_color="ff0000", \
                                              end_color="ff0000", \
                                              fill_type="solid")
            #set sheet names from dataframe columns
            sheet_names = [str(x).replace('_', ' ') for x in scoring_difference.columns]

            for subject in scoring_difference.columns:
                for sht in sheet_names:
                    sheet = scorer2_xl[sht]
                    if subject == sht.replace(' ', '_'):
                        #copy scorer1's scoring in col E
                        for line in range(self.xl_range.start_line, \
                                          self.xl_range.end_line + 1):
                            sheet.cell(row=line, \
                                       column=self.xl_range.column + 2).value = \
                                self.scorer1_scoring.loc[line - 6, subject]
                            #add scorer1's name
                            if line == self.xl_range.start_line:
                                sheet.cell(row=line - 1, \
                                           column=self.xl_range.column + 2)\
                                    .value = self.scorer1_name
                            #get cell and set style
                            if scoring_difference.loc[line - 6, subject]:
                                sheet.cell(row=line, \
                                           column=self.xl_range.column - 1)\
                                    .fill = bg_style
                        break
            scorer2_xl.save(self.scorer2_file[:-5] + '_highlighted.xlsx')
            scorer2_xl.close()
        else:
            print('Error while dropping single columns !\n' + \
                  'Highlighting of scoring differences aborted.')



class WeightedKappaFrom2Lists(WeightedCohenKappaScore):
    """ WeightedKappaFrom2Lists
        scoring1, scoring2 : Labels assigned by the first annotator and
        by the second annotator.
        The kappa statistic is symmetric,
        so swapping "scoring1" and "scoring2" doesn’t change the value.

        scorer1_name='Scorer1', scorer2_name='Scorer2' : optional

        scoring_items=None : if None, all labels that appears at least once in
        "scoring1" or "scoring2" are used.

        Example :
        import class_kappa_cohen_score as kcs
        data = pd.read_excel("./scoring.xlsx", sheet_name='sheet1').dropna()

        kappa_score = \\ \n
            kcs.WeightedKappaFrom2Lists(scoring1=data['Scorer1'], \\ \n
                                        scoring2=data['Scorer2'])\\ \n
                                        .compute_kappa()

        RESULT OUTPUT (Pandas DataFrame) :
                                                    Kappa Scores     Kappa Interpretation
        Kappa score (linear coefficients)             0.837893     Almost perfect agreement
        Kappa score (quadratic coefficients)          0.921812     Almost perfect agreement
        Kappa score (linear inverse coefficients)    -0.280056  Less than chance agreeement
        Kappa score (quadratic inverse coefficients) -0.103552  Less than chance agreeement


        krippendorff_alpha = kappa.compute_krippendorff_alpha()

        RESULT OUTPUT (Pandas Series) :
        Krippendorff alpha score                 0.933487
        Alpha score Agreement       Substantial agreement
        """


    def __init__(self, scoring1, scoring2, scorer1_name='Scorer1',  \
                 scorer2_name='Scorer2', scoring_items=None):
        """ init """
        WeightedCohenKappaScore.__init__(self)

        self.scorer1_name = scorer1_name
        self.scorer2_name = scorer2_name

        del self.xl_range
        del self.df_global
        del self.scorer2_file
        del self.result_filename
        del self.df_results

        self.scorer1_scoring = pd.Series(scoring1)
        self.scorer2_scoring = pd.Series(scoring2)

        if scoring_items is not None:
            self.scoring_items = scoring_items
        else:
            self.scoring_items = self.retrieve_kappa_items()

        #linear coefficients matrices
        self.matrix_df_linear, self.inverse_matrix_df_linear = \
            self.create_coeff_matrices(items=self.scoring_items, \
                                       weights='linear')
        #quadratic coefficents matrices
        self.matrix_df_square, self.inverse_matrix_df_square = \
            self.create_coeff_matrices(items=self.scoring_items, \
                                       weights='quadratic')

        self.kappa_score = pd.DataFrame()
        self.alpha_score = pd.Series()



    def retrieve_kappa_items(self):
        """retrieve_kappa_items"""
        import copy
        try:
            itm1 = [int(x) for x in list(set(self.scorer1_scoring))]
        except ValueError:
            itm1 = [str(x) for x in list(set(self.scorer1_scoring))]
        try:
            itm2 = [int(x) for x in list(set(self.scorer2_scoring))]
        except ValueError:
            itm2 = [str(x) for x in list(set(self.scorer2_scoring))]
        if len(itm1) < len(itm2):
            itm = copy.deepcopy(itm2)
            itm.extend([x for x in itm1 if x not in itm])
        else:
            itm = copy.deepcopy(itm1)
            itm.extend([x for x in itm2 if x not in itm])
        return itm



    def compute_kappa(self):
        """compute_single_item_kappa"""
        kappa_result = pd.DataFrame()
        kappa = pd.Series()
        kappa['Kappa score (linear coefficients)'] = \
            self.compute_kappa_score(scorer1=self.scorer1_scoring, \
                                     scorer2=self.scorer2_scoring,
                                     items=self.scoring_items,
                                     matrix_weights=self.matrix_df_linear)

        kappa['Kappa score (quadratic coefficients)'] = \
            self.compute_kappa_score(scorer1=self.scorer1_scoring, scorer2=self.scorer2_scoring,
                                     items=self.scoring_items,
                                     matrix_weights=self.matrix_df_square)

        kappa['Kappa score (linear inverse coefficients)'] = \
            self.compute_kappa_score(scorer1=self.scorer1_scoring, scorer2=self.scorer2_scoring,
                                     items=self.scoring_items,
                                     matrix_weights=self.inverse_matrix_df_linear)

        kappa['Kappa score (quadratic inverse coefficients)'] = \
            self.compute_kappa_score(scorer1=self.scorer1_scoring, scorer2=self.scorer2_scoring,
                                     items=self.scoring_items,
                                     matrix_weights=self.inverse_matrix_df_square)

        kappa_result['Kappa Scores'] = kappa
        kappa_result['Kappa Interpretation'] = self.kappa_interpretation(kappa)
        self.kappa_score = kappa_result

        return kappa_result



    def compute_krippendorff_alpha(self, level_of_measurement='ordinal'):
        """compute_krippendorff_alpha
        level_of_measurement : string or callable
        Steven's level of measurement of the variable.
        It must be one of 'nominal', 'ordinal', 'interval',
        'ratio' or a callable."""
        comp_df = pd.DataFrame(columns=[self.scorer1_name, \
                                        self.scorer2_name])
        comp_df[self.scorer1_name] = self.scorer1_scoring
        comp_df[self.scorer2_name] = self.scorer2_scoring
        comp_df = comp_df.T
        kri_res = krippendorff.alpha(reliability_data=comp_df, \
                                     value_domain=self.scoring_items, \
                                     level_of_measurement=level_of_measurement)
        self.alpha_score["Krippendorff alpha score"] = kri_res
        self.alpha_score["Alpha score Agreement"] = \
            self.alpha_interpretation(kri_res)
        return self.alpha_score



    def alpha_interpretation(self, alpha_score):
        """kripendorff alpha interpretation
           NOT SURE OF THE SCALE ! """
        if alpha_score <= 0.667:
            alpha_interpretation = 'Unreliable agreement'
        elif 0.667 < alpha_score < 0.81:
            alpha_interpretation = 'Acceptable agreement'
        elif 0.80 < alpha_score <= 1:
            alpha_interpretation = 'Substantial agreement'
        elif alpha_score == 1:
            alpha_interpretation = 'Perfect agreement'
        return alpha_interpretation


    #Overridding base class methods not compatible with new derived class
    def import_scoring_from_1_file(self):
        """import_scoring_from_1_file"""
        print('Not allowed')

    def import_scoring_from_many_files(self):
        """import_scoring_from_many_files"""
        print('Not allowed')

    def _sort_dataframe(self, dataframe):
        """sort dataframe columns by name"""
        print('Not allowed')

    def compute_global_kappa(self):
        """compute_global_kappa"""
        print('Not allowed')

    def save_results_to_excel(self):
        """save_results_to_excel"""
        print('Not allowed')

    def _drop_single_columns(self):
        """drop_single_columns"""
        print('Not allowed')

    def highlight_differences_scorer2(self):
        """highlight_differences_scorer2"""
        print('Not allowed')
